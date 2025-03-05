import json
import os

import boto3
from botocore.exceptions import NoCredentialsError, ProfileNotFound, ClientError
from openpyxl import Workbook, load_workbook


try:
    # change the profile name per your ~/.aws/credentials or ~/.aws/config
    session = boto3.Session(profile_name="default", region_name="us-east-1")
except (NoCredentialsError, ProfileNotFound):
    print('AWS credentials not found or improperly configured.')
except Exception:
    print('Error with setting up AWS session.')


def get_scuto_from_workload_estimate(workload_estimate_id):
    """
    Retrieves Service Code, Usage Type, Operation (SCUTO) values
    for each usage line from a Workload Estimate.

    Args:
        workload_estimate_id (str): The UUID of the Workload estimate.
        Can be found from AWS Pricing Calculator Console.
    """
    client_bcm_pricing_calculator = session.client("bcm-pricing-calculator")
    paginator = client_bcm_pricing_calculator.get_paginator("list_workload_estimate_usage")
    next_token = None

    try:
        while True:
            response_iterator = paginator.paginate(
                workloadEstimateId=workload_estimate_id,
                PaginationConfig={"MaxItems": 100, "PageSize": 1, "StartingToken": next_token},
            )

            for response in response_iterator:
                for usage_line in response.get("items", []):
                    get_sku_rate_code_from_scuto(
                        usage_line["serviceCode"], usage_line["usageType"], usage_line["operation"]
                    )

                next_token = response.get("nextToken")

            if not next_token:
                break

    except ClientError as e:
        if e.response['Error']['Code'] == 'ResourceNotFoundException':
            print('Workload estimate id is invalid.')
        elif e.response['Error']['Code'] == 'AccessDeniedException':
            print('You do not have access to the workload estimate.')
        elif e.response['Error']['Code'] == 'DataUnavailableException':
            print('Workload estimate is unavailable.')


def get_sku_rate_code_from_scuto(service_code, usage_type, operation):
    """
    Retrieves rate code from Price List API for the given service code, usage type, and operation.

    Args:
        service_code(str): Service code of a usage line. Example: AmazonDynamoDB
        usage_type(str): Usage type of a usage line. Example: AFS1-WriteRequestUnits
        operation(str): Operation of a usage line. Example: PayPerRequestThroughput
    """
    client_price_list_api = session.client("pricing")  # Price List API
    paginator = client_price_list_api.get_paginator("get_products")
    next_token = None

    filters = [{"Type": "TERM_MATCH", "Field": "usageType", "Value": usage_type}]

    if operation != "":
        filters.append({"Type": "TERM_MATCH", "Field": "operation", "Value": operation})

    while True:
        response_iterator = paginator.paginate(
            ServiceCode=service_code,
            Filters=filters,
            FormatVersion="aws_v1",
            PaginationConfig={"MaxItems": 5, "PageSize": 5, "StartingToken": next_token},
        )

        for response in response_iterator:
            for pricing_line in response.get("PriceList", []):
                data = json.loads(pricing_line)
                price_dimensions = data["terms"]["OnDemand"]
                print(
                    f"Service Code: {service_code}, UsageType: {usage_type}, "
                    f"Operation: {operation}"
                )
                parse_nested_json_to_xlsx(service_code, usage_type, operation, price_dimensions)

            next_token = response.get("nextToken")

        if not next_token:
            break


def parse_nested_json_to_xlsx(sc, ut, op, price_dimensions, parent_key="", rate_code=""):
    """
    Parses the nested {terms}.{OnDemand} Price List API structure to retrieve rateCode.

    Args:
        sc(str): Service code.
        ut(str): Usage type.
        op(str): Operation.
        price_dimensions(dict): {terms}.{OnDemand} Price List API structure from get_products.
    """
    check_if_excel_exists()
    target_key_rc = "rateCode"
    target_key_desc = "description"

    if isinstance(price_dimensions, dict):
        for key, value in price_dimensions.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if rate_code == "" or rate_code is None:
                rate_code = price_dimensions.get(target_key_rc)
                description = price_dimensions.get(target_key_desc)
                if rate_code:
                    print(rate_code)
                    wb = load_workbook("data.xlsx")
                    ws = wb["Sheet1"]
                    ws.append([sc, ut, op, rate_code, description])
                    wb.save("data.xlsx")

                    return
            parse_nested_json_to_xlsx(sc, ut, op, value, new_key, rate_code)


def check_if_excel_exists():
    file_path = "data.xlsx"

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Service Code", "Usage Type", "Operation", "Rate Code", "Description"])
        wb.save(file_path)


if __name__ == "__main__":
    workload_estimate_id = input("Enter Workload Estimate id: ")
    get_scuto_from_workload_estimate(workload_estimate_id)
